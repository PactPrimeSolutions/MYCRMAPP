from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.pdfgen import canvas
import io
import logging
from .models import BusinessDetails
from .forms import BusinessDetailsForm
from django.http import JsonResponse
from .models import Event
from .models import Client
from .forms import ClientForm
from openpyxl.styles import NamedStyle
from django.shortcuts import redirect
from django.utils.dateparse import parse_date

logger = logging.getLogger(__name__)

from django.http import JsonResponse
from .models import County

def base_view(request):
    """ View to handle the display and creation of business details """
    if request.method == "POST":
        form = BusinessDetailsForm(request.POST)
        if form.is_valid():
            form.save()
            logger.info('BusinessDetails saved successfully')
            messages.success(request, "Business details saved successfully!")
            return redirect('base')
        else:
            messages.error(request, "Please correct the errors below.")
            logger.error(f'Form errors: {form.errors}')
    else:
        form = BusinessDetailsForm()

    base = BusinessDetails.objects.filter(is_deleted=False)
    return render(request, 'myapp/base.html', {'form': form, 'base': base})

def delete_business_detail(request, pk):
    """ View to soft delete a business detail """
    business_detail = get_object_or_404(BusinessDetails, pk=pk)
    if not business_detail.is_deleted:
        business_detail.is_deleted = True
        business_detail.save()
        messages.success(request, "Business detail removed successfully!")
        logger.info(f'BusinessDetails with ID {pk} marked as deleted')
    else:
        messages.warning(request, "This item is already deleted.")
    return redirect('base')

def retrieve_deleted_base(request):
    """ View to list all deleted business details """
    deleted_base = BusinessDetails.objects.filter(is_deleted=True)
    return render(request, 'myapp/retrieve_deleted_business.html', {'deleted_base': deleted_base})

def restore_business_detail(request, pk):
    """ View to restore a soft deleted business detail """
    business_detail = get_object_or_404(BusinessDetails, pk=pk)
    if business_detail.is_deleted:
        business_detail.is_deleted = False
        business_detail.save()
        messages.success(request, "Business detail restored successfully!")
        logger.info(f'BusinessDetails with ID {pk} restored')
    else:
        messages.warning(request, "This item is not marked as deleted.")
    return redirect('retrieve_deleted_business')

def add_business_detail(request):
    if request.method == 'POST':
        form = BusinessDetailsForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('business_details')  # Redirect to the appropriate view after saving
        else:
            # The form has errors, so it will be rendered again with error messages
            return render(request, 'business_details/add.html', {'form': form})
    else:
        form = BusinessDetailsForm()
        return render(request, 'business_details/add.html', {'form': form})


def download_excel(request):
    """ View to export business details to an Excel file """
    wb = Workbook()
    ws = wb.active
    ws.title = "Business Details"

    # Define headers
    headers = [
        'SL', 'Batch Type', 'Order No', 'Received On', 'Borrower Name 1',
        'Borrower Name 2', 'Address', 'State', 'County', 'Loan Amount',
        'Product', 'Status', 'Processor Name', 'Typing', 'Completed On', 'Order', 'Qcer'
    ]
    ws.append(headers)

    # Define a currency style
    currency_style = NamedStyle(name='currency_style', number_format='$#,##0.00')

    # Retrieve the download option from the POST request
    download_option = request.POST.get('download_option')
    year = request.POST.get('year')
    month = request.POST.get('month')
    start_date = request.POST.get('start_date')
    end_date = request.POST.get('end_date')

    # Retrieve and filter business details data based on the selected option
    try:
        base = BusinessDetails.objects.filter(is_deleted=False)

        if download_option == 'year' and year:
            base = base.filter(received_on__year=year)

        elif download_option == 'month' and month:
            year, month = map(int, month.split('-'))
            base = base.filter(received_on__year=year, received_on__month=month)

        elif download_option == 'date_range' and start_date and end_date:
            start_date = parse_date(start_date)
            end_date = parse_date(end_date)
            base = base.filter(received_on__range=[start_date, end_date])

        for business in base:
            row = [
                business.sl, business.batch_type, business.order_no, business.received_on,
                business.borrower_name_1, business.borrower_name_2, business.address,
                business.state, business.county, business.loan_amount, business.product,
                business.status, business.processor_name, business.typing, business.completed_on,
                business.order, business.Qcer
            ]
            ws.append(row)
            # Apply currency format to the loan_amount column
            ws.cell(row=ws.max_row, column=headers.index('Loan Amount') + 1).style = currency_style

    except Exception as e:
        logger.error(f"Error while exporting to Excel: {e}")
        messages.error(request, "There was an error exporting the data to Excel.")
        return redirect('base')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=business_details_{download_option}.xlsx'
    wb.save(response)

    return response

def generate_report(request):
    """ View to generate a PDF report of business details """
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer)

    # Example content
    p.drawString(100, 750, "Business Details Report")
    p.drawString(100, 730, "This is a sample report.")

    # Example content addition
    # Format loan amount as currency
    loan_amount = 1234.56  # Example amount, replace with actual value
    p.drawString(100, 700, f"Loan Amount: ${loan_amount:,.2f}")

    p.showPage()
    p.save()

    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')

def retrieve_deleted(request):
    """ View to list all deleted business details """
    deleted_base = BusinessDetails.objects.filter(is_deleted=True)
    return render(request, 'myapp/retrieve_deleted_business.html', {'deleted_base': deleted_base})

def calendar_view(request):
    return render(request, 'myapp/calendar.html')

def events_view(request):
    events = Event.objects.all()
    events_list = [{
        'title': event.title,
        'start': event.start.isoformat(),
        'end': event.end.isoformat() if event.end else None,
        'description': event.description,
    } for event in events]

    return JsonResponse(events_list, safe=False)

def event_list(request):
    events = Event.objects.all()
    events_data = []
    for event in events:
        events_data.append({
            'title': event.title,
            'start': event.start_time.strftime('%Y-%m-%dT%H:%M:%S'),
            'end': event.end_time.strftime('%Y-%m-%dT%H:%M:%S'),
        })
    return JsonResponse(events_data, safe=False)

def delete_event(request, event_id):
    event = CalendarEvent.objects.get(id=event_id)
    event.is_deleted = True
    event.save()
    return redirect('calendar')

def restore_event(request, event_id):
    event = CalendarEvent.objects.get(id=event_id)
    event.is_deleted = False
    event.save()
    return redirect('calendar')

def add_event(request):
    if request.method == 'POST':
        title = request.POST['title']
        description = request.POST.get('description', '')
        start_time = request.POST['start_time']
        end_time = request.POST['end_time']
        CalendarEvent.objects.create(
            title=title,
            description=description,
            start_time=start_time,
            end_time=end_time
        )
        return redirect('calendar')
    return render(request, 'add_event.html')





def client_list_view(request):
    clients = Client.objects.filter(deleted=False)
    return render(request, 'myapp/client_list.html', {'clients': clients})

def client_create_view(request):
    if request.method == "POST":
        form = ClientForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Client details saved successfully!")
            return redirect('client_list')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = ClientForm()

    return render(request, 'myapp/client_form.html', {'form': form})

def client_edit_view(request, pk):
    client = get_object_or_404(Client, pk=pk)
    if request.method == "POST":
        form = ClientForm(request.POST, instance=client)
        if form.is_valid():
            form.save()
            messages.success(request, "Client details updated successfully!")
            return redirect('client_list')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = ClientForm(instance=client)

    return render(request, 'myapp/client_form.html', {'form': form})

def client_delete_view(request, pk):
    client = get_object_or_404(Client, pk=pk)
    if request.method == "POST":  # Confirm deletion
        client.deleted = True
        client.save()
        messages.success(request, "Client deleted successfully!")
        return redirect('client_list')
    return render(request, 'myapp/confirm_delete.html', {'client': client})

def client_detail_view(request, pk):
    client = get_object_or_404(Client, pk=pk)
    return render(request, 'myapp/client_detail.html', {'client': client})


def add_client_view(request):
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('client_list')
    else:
        form = ClientForm()
    return render(request, 'myapp/add_client.html', {'form': form})


def delete_client_view(request, pk):
    client = get_object_or_404(Client, pk=pk)
    client.deleted = True
    client.save()
    return redirect('client_list')

def deleted_clients_view(request):
    deleted_clients = Client.objects.filter(deleted=True)
    return render(request, 'myapp/deleted_clients.html', {'deleted_clients': deleted_clients})

def restore_client_view(request, pk):
    client = get_object_or_404(Client, pk=pk)
    client.deleted = False
    client.save()
    return redirect('client_list')




def load_counties(request):
    state_id = request.GET.get('state_id')
    counties = County.objects.filter(state_id=state_id).order_by('name')
    return JsonResponse(list(counties.values('id', 'name')), safe=False) 


def get_counties(request, state_id):
    counties = County.objects.filter(state_id=state_id).values('id', 'name')
    return JsonResponse(list(counties), safe=False)